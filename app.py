import pandas as pd
import openpyxl
from openpyxl.styles import Font

#Lista vazia de alunos
alunos = []

while True:
    nome = str(input("Digite um nome: ")).strip().title()
    nota1 = float(input("Nota 1: "))
    nota2 = float(input("Nota 2: "))
    nota3 = float(input("Nota 3: "))
    nota4 = float(input("Nota 4: "))

    media = (nota1 + nota2 + nota3 + nota4) / 4
    situacao = "Aprovado(a)" if media >= 7 else "Reprovado(a)"

#guardar dados dos alunos em um dicionario

    aluno = {
        "Nome": nome,
        "Media": media,
        "Situacao": situacao
    }

    #adiciona aluno a lista
    alunos.append(aluno)

    # pergunta se quer continuar:
    continuar = input("\nQuer adicionar outro aluno ? (S/N): ").upper()
    if continuar != 'S':
        break

tabela_alunos = pd.DataFrame(alunos)

#criando planilha
book = openpyxl.Workbook()

#criando uma pagina
pag_alunos = book.active
pag_alunos.title = "Alunos"

#adicionando cabeçalho
pag_alunos.append(["Nome", "Média", "Situação"])



#adicionar dados a planilha
for aluno in alunos:
    pag_alunos.append([aluno["Nome"], aluno["Media"], aluno["Situacao"]])
    
    #aplicando cor base na situacao
    if aluno["Situacao"] == "Aprovado(a)":
        pag_alunos.cell(row=pag_alunos.max_row, column=3).font = Font(color="00FF00")
    else:
        pag_alunos.cell(row=pag_alunos.max_row, column=3).font = Font(color="FF0000")

#salvar planilha
book.save("planilha de alunos.xlsx")